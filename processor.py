
import os, json
from uuid import uuid4
import pandas as pd
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

ALIAS = [
    ("Mohammed Salih","Mohammaed Salih"),
    ("Michael Molley","Michael Molley Jr"),("Michael Molley","Michael MolleyJr"),
    ("Jalyn Folston","Jayln Folston"),("Darrell Clark","Darrell jr Clark"),
    ("Aminata Jallow","Aminata Jalllow"),("Haneef Naseem","Haneef Nassem"),
    ("Pierre Patrick Julien","Peirre Julien"),("Quincy Jones","Quincy jermain Jones"),
    ("Briante Guillory","Briante GuilloryMelvin"),("Francisco Amsa","Franscisco Amsa"),
    ("Kiarra Bonilla","Kiko McQueen"),("Devin Smith","Devin Traion Smith"),
    ("JianYangYu","Jian Yu"),("Awunglefac Nzeffe","Awung Zeff"),
    ("Indiya Gray","Indiya Jahnise Gray"),("TyNia Andrews","Ty-Nia Andrews"),
    ("Bri Johnson","Brianna Johnson"),("Devante Young","Devonte Young"),
    ("Saleem Westfield","Salem Westfield"),("Timothy Gant jr","Timothy Gant"),
    ("McQueen, Kiko","Kiarra,Bonilla"),("Gant, Timothy","Timothy,Gant jr"),
    ("Smith, Devin Traion","Devin,Smith"),("Jones, Quincy","Quincy jermain,Jones"),
]

def norm(s:str)->str:
    if not isinstance(s,str): return ''
    return s.lower().replace(',','').replace(' ','').replace('-','').replace("'",'').replace('.','')

def split_name(s:str):
    if not isinstance(s,str): return ('','')
    if ',' in s:
        parts = [p.strip() for p in s.split(',',1)]
        return (parts[1], parts[0]) if len(parts)==2 else (s,'')
    ps = s.split()
    return (ps[0], ps[-1]) if len(ps)>=2 else (s,'')

def keys_for(s:str):
    f,l = split_name(s)
    out = {
        norm(s),
        norm(f+','+l),
        norm(l+','+f),
        norm(f+l), norm(l+f)
    }
    if f and l:
        out.add(norm(f[0]+l)); out.add(norm(l+f[0]))
    amap = {norm(b): norm(a) for a,b in ALIAS}
    amap.update({norm(a): norm(a) for a,b in ALIAS})
    raw = norm(s)
    if raw in amap: out.add(amap[raw])
    return list(out)

def to_hhmmss(hours):
    if pd.isna(hours): return None
    total = int(round(float(hours)*3600))
    sign = '-' if total<0 else ''
    total = abs(total)
    h, r = divmod(total,3600); m, s = divmod(r,60)
    return f"{sign}{h:02d}:{m:02d}:{s:02d}"

def process_files(cortex_path: str, adp_path: str, out_dir: str):
    # Cortex
    C = pd.read_excel(cortex_path)
    C.columns = [c.strip().lower().replace(' ','_') for c in C.columns]
    C['app_sign_in:']  = pd.to_datetime(C['app_sign_in:'], errors='coerce')
    C['app_sign_out:'] = pd.to_datetime(C['app_sign_out:'], errors='coerce')
    C['cortex_total_break_time_used'] = pd.to_numeric(C['cortex_total_break_time_used'], errors='coerce').fillna(0)
    C['total_cortex'] = (C['app_sign_out:'] - C['app_sign_in:']).dt.total_seconds()/3600
    C['net_cortex']   = C['total_cortex'] - (C['cortex_total_break_time_used']/60)
    Cgrp = C.groupby('driver_name', as_index=False).agg(
        **{
            'sum_break_minutes': ('cortex_total_break_time_used','sum'),
            'total_cortex': ('total_cortex','sum'),
            'net_cortex': ('net_cortex','sum'),
        }
    )

    # ADP (First + Last -> "Last, First"), or accept complete_name
    A = pd.read_csv(adp_path)
    A.columns = [c.strip().lower().replace(' ','_') for c in A.columns]
    if 'first_name' in A.columns and 'last_name' in A.columns:
        A['driver_name'] = A['last_name'].astype(str).str.strip()+', '+A['first_name'].astype(str).str.strip()
    elif 'complete_name' in A.columns:
        A['driver_name'] = A['complete_name']
    else:
        raise ValueError('ADP file missing First/Last or complete name columns')

    if 'in_time' in A.columns: A['in_time'] = pd.to_datetime(A['in_time'], errors='coerce')
    if 'out_time' in A.columns: A['out_time'] = pd.to_datetime(A['out_time'], errors='coerce')
    if 'hours' in A.columns:
        A['hours'] = pd.to_numeric(A['hours'], errors='coerce')
        Ahours = A.groupby('driver_name', as_index=False)['hours'].sum().rename(columns={'hours':'total_hours_worked'})
    else:
        A['duration'] = (A['out_time']-A['in_time']).dt.total_seconds()/3600
        Ahours = A.groupby('driver_name', as_index=False)['duration'].sum().rename(columns={'duration':'total_hours_worked'})

    A_sorted = A.sort_values(['driver_name','in_time'])
    A_sorted['break_mins'] = (A_sorted['in_time'] - A_sorted['out_time'].shift()).dt.total_seconds()/60
    A_sorted['break_mins'] = A_sorted['break_mins'].apply(lambda x: x if (pd.notna(x) and x>0) else 0)
    Abreaks = A_sorted.groupby('driver_name', as_index=False)['break_mins'].sum().rename(columns={'break_mins':'break_minutes'})

    Agrp = pd.merge(Ahours, Abreaks, on='driver_name', how='outer')

    # Two-way matching
    Ckeys = []
    for i, r in Cgrp.iterrows():
        for k in keys_for(r['driver_name']):
            Ckeys.append({'c_idx':i,'key':k})
    Ckeys = pd.DataFrame(Ckeys)

    Akeys = []
    for i, r in Agrp.iterrows():
        for k in keys_for(r['driver_name']):
            Akeys.append({'a_idx':i,'key':k})
    Akeys = pd.DataFrame(Akeys)

    cand = pd.merge(Ckeys, Akeys, on='key', how='inner').drop_duplicates(['c_idx','a_idx'])

    matches = []
    for c_idx, g in cand.groupby('c_idx'):
        c_hours = Cgrp.loc[c_idx,'net_cortex']
        if len(g)==1:
            matches.append((c_idx, int(g.iloc[0]['a_idx'])))
        else:
            best=None; bestdiff=None
            for _, row in g.iterrows():
                a_idx = int(row['a_idx'])
                adp_h = Agrp.loc[a_idx,'total_hours_worked']
                diff = abs(c_hours - adp_h) if (pd.notna(c_hours) and pd.notna(adp_h)) else 9999
                if best is None or diff<bestdiff:
                    best, bestdiff = a_idx, diff
            matches.append((c_idx, best))

    mDF = pd.DataFrame(matches, columns=['c_idx','a_idx']).drop_duplicates()

    C_only = Cgrp[~Cgrp.index.isin(mDF['c_idx'])]
    A_only = Agrp[~Agrp.index.isin(mDF['a_idx'])]

    rows = []
    for _, r in mDF.iterrows():
        Cx = Cgrp.loc[r['c_idx']]
        Ax = Agrp.loc[r['a_idx']]
        rows.append({
            'Cortex Name': Cx['driver_name'],
            'ADP Name': Ax['driver_name'],
            'Sum of Break Duration in Minutes': round(Cx['sum_break_minutes'],2),
            'Total Cortex': round(Cx['total_cortex'],2),
            'Net Cortex': round(Cx['net_cortex'],2),
            'ADP (Decimal)': round(Ax['total_hours_worked'],2),
        })
    for _, Cx in C_only.iterrows():
        rows.append({
            'Cortex Name': Cx['driver_name'], 'ADP Name': None,
            'Sum of Break Duration in Minutes': round(Cx['sum_break_minutes'],2),
            'Total Cortex': round(Cx['total_cortex'],2), 'Net Cortex': round(Cx['net_cortex'],2),
            'ADP (Decimal)': None,
        })
    for _, Ax in A_only.iterrows():
        rows.append({
            'Cortex Name': None, 'ADP Name': Ax['driver_name'],
            'Sum of Break Duration in Minutes': None, 'Total Cortex': None, 'Net Cortex': None,
            'ADP (Decimal)': round(Ax['total_hours_worked'],2),
        })

    F = pd.DataFrame(rows)
    F['ADP (hh:mm:ss)'] = F['ADP (Decimal)'].apply(lambda x: to_hhmmss(x) if pd.notna(x) else None)
    F['Net Cortex (hh:mm:ss)'] = F['Net Cortex'].apply(lambda x: to_hhmmss(x) if pd.notna(x) else None)

    def td_from_hours(x):
        return timedelta(hours=float(x)) if pd.notna(x) else None

    def diff_hhmmss(adp_dec, cortex_dec):
        if pd.isna(adp_dec) or pd.isna(cortex_dec):
            return None
        td = td_from_hours(adp_dec) - td_from_hours(cortex_dec)
        total = int(td.total_seconds())
        sign = '-' if total<0 else ''
        total = abs(total)
        h, r = divmod(total,3600); m, s = divmod(r,60)
        return f"{sign}{h:02d}:{m:02d}:{s:02d}"

    F['Difference (hh:mm:ss)'] = F.apply(lambda r: diff_hhmmss(r['ADP (Decimal)'], r['Net Cortex']), axis=1)

    # Excel + formatting
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook(); ws = wb.active; ws.title = 'Cortex vs ADP'
    for row in dataframe_to_rows(F[[
        'Cortex Name','ADP Name','Sum of Break Duration in Minutes','Total Cortex',
        'Net Cortex','Net Cortex (hh:mm:ss)','ADP (hh:mm:ss)','ADP (Decimal)','Difference (hh:mm:ss)'
    ]], index=False, header=True):
        ws.append(row)

    Y = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    R = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')

    headers = [cell.value for cell in ws[1]]
    col_cortex = headers.index('Cortex Name')+1
    col_adp    = headers.index('ADP Name')+1
    col_diff   = headers.index('Difference (hh:mm:ss)')+1

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cortex_name = row[col_cortex-1].value
        adp_name    = row[col_adp-1].value
        diffv       = row[col_diff-1].value
        if not cortex_name or not adp_name:
            for c in row: c.fill = Y
        if isinstance(diffv, str) and diffv.startswith('-'):
            row[col_diff-1].fill = R

    os.makedirs(out_dir, exist_ok=True)
    excel_path = os.path.join(out_dir, f"Reconciliation_{uuid4().hex}.xlsx")
    wb.save(excel_path)

    # Summary sections
    A = A  # keep scope
    Agrp = Agrp
    Cgrp = Cgrp

    A_over11 = Agrp[(Agrp['total_hours_worked']>11)]['driver_name'].dropna().tolist()
    C_lt30 = Cgrp[(Cgrp['sum_break_minutes']<30)]['driver_name'].dropna().tolist()
    c_only = F[F['Cortex Name'].notna() & F['ADP Name'].isna()]['Cortex Name'].dropna().tolist()
    a_only = F[F['ADP Name'].notna() & F['Cortex Name'].isna()]['ADP Name'].dropna().tolist()
    neg_rows = F[F['Difference (hh:mm:ss)'].astype(str).str.startswith('-')]
    neg_names = [(r['ADP Name'] if pd.notna(r['ADP Name']) else r['Cortex Name']) for _, r in neg_rows.iterrows() if (pd.notna(r['ADP Name']) or pd.notna(r['Cortex Name']))]

    lines = []
    lines.append("Cortex vs ADP Summary")
    lines.append("")
    lines.append("Worked > 11 Hours")
    lines += [f"- {n}" for n in A_over11] or ["- None"]
    lines.append("")
    lines.append("Took < 30 Minutes Break")
    lines += [f"- {n}" for n in C_lt30] or ["- None"]
    lines.append("")
    lines.append("In Cortex but NOT in ADP")
    lines += [f"- {n}" for n in c_only] or ["- None"]
    lines.append("")
    lines.append("In ADP but NOT in Cortex")
    lines += [f"- {n}" for n in a_only] or ["- None"]
    lines.append("")
    lines.append("ADP Hours < Net Cortex Hours")
    lines += [f"- {n}" for n in neg_names] or ["- None"]

    return {"excel_file": excel_path, "summary": "\n".join(lines)}
